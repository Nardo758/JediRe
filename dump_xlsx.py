import openpyxl, sys, json
wb = openpyxl.load_workbook('jedi_re_module_wiring_blueprint_v2.xlsx')

data = {}
for sn in wb.sheetnames:
    ws = wb[sn]
    rows = []
    for r in range(1, ws.max_row+1):
        vals = []
        for c in range(1, ws.max_column+1):
            v = ws.cell(r, c).value
            if v is None:
                vals.append('')
            else:
                s = str(v)[:100].strip().replace('\n',' ').replace('\r','').replace('→','->').replace('←','<-')
                # Filter non-ASCII
                s = ''.join(c if ord(c) < 128 else '?' for c in s)
                vals.append(s)
        rows.append(vals)
    data[sn] = rows

with open('xlsx_dump.json', 'w') as f:
    json.dump(data, f, indent=1)
print('Done')
